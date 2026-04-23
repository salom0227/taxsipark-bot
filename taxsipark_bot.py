"""
🚖 Taksopark Bot — To'liq versiya
  • Admin panel (tugmalar)
  • Guruhga bildirishnoma (start / ro'yxat)
  • Broadcast (hammaga / ro'yxatdagilarga / start bosganlar)
  • Excel eksport
  • Reklama rejalashtirish
"""

import logging
import os
import asyncio
import io
import httpx
from datetime import datetime, timezone
from contextlib import asynccontextmanager
from zoneinfo import ZoneInfo

from fastapi import FastAPI, Request
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)
from telegram.request import HTTPXRequest

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

from database import (
    init_db, close_pool,
    save_user, get_all_users, get_all_user_ids,
    get_users_with_time, count_users,
    log_login,
    save_start, get_all_starts, get_all_start_ids, count_starts,
    add_scheduled_ad, get_pending_ads, get_all_scheduled_ads,
    mark_ad_sent, delete_scheduled_ad,
)

# ─── CONFIG ──────────────────────────────────────────────────────────────────

BOT_TOKEN       = os.environ.get("BOT_TOKEN", "8670099128:AAEaLw1r4GmoVHPOjgk8NXCKJQbksxY5-co")
ADMIN_USERNAMES = {"SAFARGO_TAXI", "salom0227", "ibrokhim_515", "Fixonee"}
ADMIN_IDS       = [5567499156, 659123909, 8070344459]
GROUP_ID        = int(os.environ.get("GROUP_ID", "-5234833498"))
RENDER_URL      = os.environ.get("RENDER_URL", "https://taxsipark-bot.onrender.com")
WELCOME_IMAGE   = "welcome.png"
WEBHOOK_PATH    = f"/webhook/{BOT_TOKEN}"
WEBHOOK_URL     = f"{RENDER_URL}{WEBHOOK_PATH}"
TZ              = ZoneInfo("Asia/Tashkent")

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ─── STATES ──────────────────────────────────────────────────────────────────

REG_NAME, REG_PHONE, MAIN_MENU = range(3)
BROADCAST_TEXT, BROADCAST_TARGET = range(10, 12)
AD_TEXT, AD_TARGET, AD_DATETIME = range(20, 23)
EDIT_KEY, EDIT_VALUE = range(30, 32)

# ─── MATNLAR (tahrir qilinadigan) ─────────────────────────────────────────────

TEXTS = {
    "welcome": (
        "👋 Salom! 🚖 <b>Taksopark</b> botiga xush kelibsiz!\n\n"
        "Boshlash uchun <b>to'liq ismingizni</b> kiriting (F.I.Sh) 👇"
    ),
    "malumot": (
        "🚖 Taksopark — haydovchilar uchun tezkor ro'yxatdan o'tish tizimi\n"
        "🚖 Safargo — haydovchilar uchun eng qulay taksopark!\n\n"
        "Assalomu alaykum! 👋\n"
        "Safargo sizga nafaqat ish, balki barqaror daromad va bonuslar ham beradi 💸\n\n"
        "🎁 Start bonus:\n"
        "Ro'yxatdan o'tganingiz zahoti 40 000 so'm bonus sizniki!\n\n"
        "🔥 Har oy sovg'alar:\n"
        "Safargo'da har oy yangi bonus va aksiyalar bo'lib turadi 🎉\n\n"
        "👨‍👨‍👦 Do'st olib keling – pul ishlang:\n"
        "Har bir do'st uchun 50 000 so'm\n"
        "(50 ta zakaz bajargach beriladi) 💰\n\n"
        "📉 Minimal foiz:\n"
        "Atigi 2,2% — siz uchun maksimal foyda ⚡\n\n"
        "🏆 Katta balans = katta bonus:\n"
        "1 mln so'm yig'ing — 50 000 so'm bonus oling 🎯\n\n"
        "💳 Qulay to'lov:\n"
        "Click / Payme orqali to'ldiring va 5% keshbek oling 🔥\n\n"
        "📜 Ishonch va rasmiylik:\n"
        "Safargo — sertifikatlangan taksopark ✔️\n\n"
        "🚘 Qo'shimcha imkoniyatlar:\n"
        "✅ Litsenziya nakleyka\n"
        "✅ Komfort / Komfort+ ochish\n"
        "✅ Blokdan chiqarish xizmati\n\n"
        "🚀 Safargo bilan ishlash — bu o'sish, daromad va qulaylik!\n"
        "Bugunoq bizga qo'shiling 🔥"
    ),
    "boglanish": (
        "📞 Operator bilan bog'lanish:\n\n"
        "📱 Telefon: +998(55)515-00-54\n"
    ),
    "hujjat": (
        "📋 Prava va texpasportingizni yuboring adminimizga!\n\n"
        "Hujjatlaringizni rasm yoki fayl shaklida yuboring 👇"
    ),
}

# ─── KEYBOARDS ───────────────────────────────────────────────────────────────

def kb_main():
    return ReplyKeyboardMarkup([
        ["🚖 Taksopark haqida ma'lumot olish"],
        ["📞 Operator bilan bog'lanish"],
        ["📝 Ro'yxatdan o'tish"],
    ], resize_keyboard=True)

def kb_phone():
    return ReplyKeyboardMarkup(
        [[KeyboardButton("📱 Raqamni avtomatik yuborish", request_contact=True)]],
        resize_keyboard=True, one_time_keyboard=True
    )

def kb_admin_link():
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("📤 Hujjat yuborish", url="https://t.me/SAFARGO_TAXI")
    ]])

def kb_admin_panel():
    return ReplyKeyboardMarkup([
        ["📊 Statistika", "👥 Ro'yxat (Excel)"],
        ["✏️ Matnlarni tahrirlash", "📢 Broadcast"],
        ["📅 Reklama rejalashtirish", "⏰ Jadval ko'rish"],
        ["🔙 Chiqish"],
    ], resize_keyboard=True)

def kb_broadcast_target():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 Hammaga (start bosganlar)", callback_data="bc_all")],
        [InlineKeyboardButton("✅ Ro'yxatdagilarga", callback_data="bc_registered")],
        [InlineKeyboardButton("❌ Bekor qilish", callback_data="bc_cancel")],
    ])

def kb_ad_target():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 Hammaga", callback_data="ad_all")],
        [InlineKeyboardButton("✅ Ro'yxatdagilarga", callback_data="ad_registered")],
        [InlineKeyboardButton("❌ Bekor qilish", callback_data="ad_cancel")],
    ])

def kb_edit_texts():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👋 Xush kelibsiz matni", callback_data="edit_welcome")],
        [InlineKeyboardButton("🚖 Ma'lumot matni", callback_data="edit_malumot")],
        [InlineKeyboardButton("📞 Bog'lanish matni", callback_data="edit_boglanish")],
        [InlineKeyboardButton("📋 Hujjat matni", callback_data="edit_hujjat")],
        [InlineKeyboardButton("❌ Bekor qilish", callback_data="edit_cancel")],
    ])

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def is_admin(update: Update) -> bool:
    user = update.effective_user
    if user.id in ADMIN_IDS:
        return True
    if user.username and user.username.upper() in {u.upper() for u in ADMIN_USERNAMES}:
        return True
    return False

async def notify_group(bot, text: str):
    """Guruhga xabar yuborish"""
    if GROUP_ID == 0:
        return
    try:
        await bot.send_message(chat_id=GROUP_ID, text=text, parse_mode="HTML")
    except Exception as e:
        logger.warning(f"Guruhga xabar yuborilmadi: {e}")

def make_excel_users(users: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ro'yxatdagilar"

    headers = ["#", "Ism", "Telefon", "Username", "Ro'yxat sanasi"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, u in enumerate(users, 1):
        reg = u.get("registered_at")
        reg_str = reg.astimezone(TZ).strftime("%d.%m.%Y %H:%M") if reg else "—"
        ws.append([i, u["name"], u["phone"], u["username"], reg_str])

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def make_excel_starts(starts: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Start bosganlar"

    headers = ["#", "To'liq ism", "Username", "Start vaqti"]
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, s in enumerate(starts, 1):
        st = s.get("started_at")
        st_str = st.astimezone(TZ).strftime("%d.%m.%Y %H:%M") if st else "—"
        ws.append([i, s["full_name"], s["username"], st_str])

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── USER HANDLERS ────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    context.user_data.clear()

    # Start bosganlarni saqlash
    uname = f"@{user.username}" if user.username else "—"
    full_name = user.full_name or "—"
    await save_start(user.id, uname, full_name)

    # Guruhga bildirishnoma
    await notify_group(
        context.bot,
        f"👀 <b>Start bosdi:</b> {full_name} {uname}"
    )

    caption = TEXTS["welcome"]
    try:
        if os.path.exists(WELCOME_IMAGE):
            with open(WELCOME_IMAGE, "rb") as photo:
                await update.message.reply_photo(photo=photo, caption=caption, parse_mode="HTML")
        else:
            await update.message.reply_text(caption, parse_mode="HTML")
    except Exception as e:
        logger.error(f"Rasm xatosi: {e}")
        await update.message.reply_text(caption, parse_mode="HTML")
    return REG_NAME


async def handle_reg_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["name"] = update.message.text.strip()
    await update.message.reply_text(
        "📱 Telefon raqamingizni yuboring:",
        reply_markup=kb_phone()
    )
    return REG_PHONE


async def handle_reg_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = (
        update.message.contact.phone_number
        if update.message.contact
        else update.message.text.strip()
    )
    user = update.effective_user
    name = context.user_data.get("name", "—")
    uname = f"@{user.username}" if user.username else "—"

    await save_user(user.id, name, phone, uname)
    await log_login(user.id)

    # Guruhga bildirishnoma
    await notify_group(
        context.bot,
        f"✅ <b>Ro'yxatdan o'tdi:</b>\n"
        f"👤 {name}\n"
        f"📱 {phone}\n"
        f"🔗 {uname}"
    )

    # Adminga bildirishnoma
    now = datetime.now(TZ).strftime("%d.%m.%Y %H:%M")
    notif = (
        f"🆕 <b>Yangi foydalanuvchi!</b>\n\n"
        f"👤 {name}\n📱 {phone}\n🔗 {uname}\n🕐 {now}"
    )
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(chat_id=admin_id, text=notif, parse_mode="HTML")
        except Exception:
            pass

    await update.message.reply_text(
        f"✅ <b>Ro'yxatdan o'tdingiz!</b>\n\n"
        f"👤 <b>{name}</b>\n📱 <b>{phone}</b>\n\n"
        f"Tez orada operator siz bilan bog'lanadi! 🚖",
        parse_mode="HTML",
        reply_markup=kb_main()
    )
    return MAIN_MENU


async def handle_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text

    # Admin tugmalarini conversation ichida ham ushlash
    if is_admin(update):
        admin_buttons = [
            "📊 Statistika", "👥 Ro'yxat (Excel)",
            "✏️ Matnlarni tahrirlash", "📢 Broadcast",
            "📅 Reklama rejalashtirish", "⏰ Jadval ko'rish", "🔙 Chiqish"
        ]
        if text in admin_buttons:
            await handle_admin_panel(update, context)
            return MAIN_MENU
        if context.user_data.get("state") is not None:
            await handle_broadcast_or_ad_text(update, context)
            return MAIN_MENU

    if text == "🚖 Taksopark haqida ma'lumot olish":
        await update.message.reply_text(TEXTS["malumot"], reply_markup=kb_main())
    elif text == "📞 Operator bilan bog'lanish":
        await update.message.reply_text(TEXTS["boglanish"], reply_markup=kb_main())
    elif text == "📝 Ro'yxatdan o'tish":
        await update.message.reply_text(TEXTS["hujjat"], reply_markup=kb_admin_link())
    return MAIN_MENU


async def global_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    if text == "🚖 Taksopark haqida ma'lumot olish":
        await update.message.reply_text(TEXTS["malumot"], reply_markup=kb_main())
    elif text == "📞 Operator bilan bog'lanish":
        await update.message.reply_text(TEXTS["boglanish"], reply_markup=kb_main())
    elif text == "📝 Ro'yxatdan o'tish":
        await update.message.reply_text(TEXTS["hujjat"], reply_markup=kb_admin_link())
    else:
        await update.message.reply_text("Iltimos, /start bosing.", reply_markup=kb_main())

# ─── ADMIN PANEL ─────────────────────────────────────────────────────────────

async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("⛔ Sizda ruxsat yo'q.")
        return
    await update.message.reply_text(
        "🛠 <b>Admin panel</b>\n\nQuyidagi bo'limlardan birini tanlang:",
        parse_mode="HTML",
        reply_markup=kb_admin_panel()
    )


async def handle_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        return
    text = update.message.text

    # ── Statistika ──
    if text == "📊 Statistika":
        total_starts = await count_starts()
        total_users  = await count_users()
        await update.message.reply_text(
            f"📊 <b>Statistika</b>\n\n"
            f"👀 Start bosganlar: <b>{total_starts}</b>\n"
            f"✅ Ro'yxatdan o'tganlar: <b>{total_users}</b>",
            parse_mode="HTML",
            reply_markup=kb_admin_panel()
        )

    # ── Excel ──
    elif text == "👥 Ro'yxat (Excel)":
        if not EXCEL_OK:
            await update.message.reply_text("❌ openpyxl o'rnatilmagan!")
            return
        await update.message.reply_text("⏳ Excel tayyorlanmoqda...")

        # Ro'yxatdagilar
        users = await get_users_with_time()
        buf_u = make_excel_users(users)
        await update.message.reply_document(
            document=buf_u,
            filename=f"royxatdagilar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            caption=f"✅ Ro'yxatdagilar: {len(users)} ta"
        )

        # Start bosganlar
        starts = await get_all_starts()
        buf_s = make_excel_starts(starts)
        await update.message.reply_document(
            document=buf_s,
            filename=f"start_bosganlar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            caption=f"👀 Start bosganlar: {len(starts)} ta"
        )

    # ── Matn tahrirlash ──
    elif text == "✏️ Matnlarni tahrirlash":
        await update.message.reply_text(
            "✏️ Qaysi matnni tahrirlaysiz?",
            reply_markup=kb_edit_texts()
        )

    # ── Broadcast ──
    elif text == "📢 Broadcast":
        context.user_data["state"] = "broadcast"
        await update.message.reply_text(
            "📢 <b>Broadcast</b>\n\nYuboriladigan matnni kiriting:",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup([["❌ Bekor qilish"]], resize_keyboard=True)
        )

    # ── Reklama rejalashtirish ──
    elif text == "📅 Reklama rejalashtirish":
        context.user_data["state"] = "ad_text"
        await update.message.reply_text(
            "📅 <b>Reklama rejalashtirish</b>\n\n1️⃣ Reklama matnini kiriting:",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardMarkup([["❌ Bekor qilish"]], resize_keyboard=True)
        )

    # ── Jadval ko'rish ──
    elif text == "⏰ Jadval ko'rish":
        ads = await get_all_scheduled_ads()
        if not ads:
            await update.message.reply_text(
                "📭 Rejalashtirilgan reklamalar yo'q.",
                reply_markup=kb_admin_panel()
            )
            return
        for ad in ads:
            t = ad["scheduled_at"].astimezone(TZ).strftime("%d.%m.%Y %H:%M")
            target_label = "Hammaga" if ad["target"] == "all" else "Ro'yxatdagilarga"
            await update.message.reply_text(
                f"🆔 ID: <b>{ad['id']}</b>\n"
                f"📅 Vaqt: <b>{t}</b>\n"
                f"👥 Kimga: <b>{target_label}</b>\n\n"
                f"{ad['text'][:200]}{'...' if len(ad['text'])>200 else ''}",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(f"🗑 #{ad['id']} ni bekor qil", callback_data=f"deladid_{ad['id']}")
                ]])
            )

    # ── Chiqish ──
    elif text == "🔙 Chiqish":
        context.user_data.clear()
        await update.message.reply_text(
            "↩️ Asosiy menyu",
            reply_markup=kb_main()
        )


# ─── BROADCAST FLOW ──────────────────────────────────────────────────────────

async def handle_broadcast_or_ad_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Broadcast yoki reklama matni kiritilganda"""
    if not is_admin(update):
        return
    text = update.message.text

    if text == "❌ Bekor qilish":
        context.user_data.clear()
        await update.message.reply_text("❌ Bekor qilindi.", reply_markup=kb_admin_panel())
        return

    state = context.user_data.get("state")

    if state == "broadcast":
        context.user_data["bc_text"] = text
        await update.message.reply_text(
            "📢 Kimga yuborilsin?",
            reply_markup=kb_broadcast_target()
        )
        context.user_data["state"] = "broadcast_target"

    elif state == "ad_text":
        context.user_data["ad_text"] = text
        context.user_data["state"] = "ad_target"
        await update.message.reply_text(
            "2️⃣ Kimga yuborilsin?",
            reply_markup=kb_ad_target()
        )

    elif state == "ad_datetime":
        # Sana/vaqt kiritildi: format "DD.MM.YYYY HH:MM"
        try:
            dt = datetime.strptime(text.strip(), "%d.%m.%Y %H:%M")
            dt = dt.replace(tzinfo=TZ)
            ad_text   = context.user_data.get("ad_text", "")
            ad_target = context.user_data.get("ad_target", "all")
            ad_id = await add_scheduled_ad(ad_text, ad_target, dt)
            target_label = "Hammaga" if ad_target == "all" else "Ro'yxatdagilarga"
            await update.message.reply_text(
                f"✅ Reklama rejalashtirildi!\n\n"
                f"🆔 ID: <b>{ad_id}</b>\n"
                f"📅 Vaqt: <b>{dt.strftime('%d.%m.%Y %H:%M')}</b>\n"
                f"👥 Kimga: <b>{target_label}</b>",
                parse_mode="HTML",
                reply_markup=kb_admin_panel()
            )
            context.user_data.clear()
        except ValueError:
            await update.message.reply_text(
                "❌ Noto'g'ri format! Masalan: <code>25.12.2024 14:30</code>",
                parse_mode="HTML"
            )

    elif state == "edit_waiting":
        key = context.user_data.get("edit_key")
        if key and key in TEXTS:
            TEXTS[key] = text
            await update.message.reply_text(
                f"✅ Matn yangilandi!", reply_markup=kb_admin_panel()
            )
        context.user_data.clear()


# ─── CALLBACK QUERY HANDLERS ─────────────────────────────────────────────────

async def handle_callbacks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    # ── Broadcast target ──
    if data in ("bc_all", "bc_registered", "bc_cancel"):
        if data == "bc_cancel":
            await query.edit_message_text("❌ Bekor qilindi.")
            context.user_data.clear()
            return

        target = "all" if data == "bc_all" else "registered"
        bc_text = context.user_data.get("bc_text", "")

        if target == "all":
            ids = await get_all_start_ids()
            label = "hammaga"
        else:
            ids = await get_all_user_ids()
            label = "ro'yxatdagilarga"

        await query.edit_message_text(f"⏳ {label.capitalize()} yuborilmoqda ({len(ids)} ta)...")
        ok, fail = 0, 0
        for uid in ids:
            try:
                await context.bot.send_message(chat_id=uid, text=bc_text)
                ok += 1
                await asyncio.sleep(0.05)
            except Exception:
                fail += 1

        await query.message.reply_text(
            f"✅ Broadcast tugadi!\n✅ Yuborildi: {ok}\n❌ Xato: {fail}",
            reply_markup=kb_admin_panel()
        )
        context.user_data.clear()

    # ── Ad target ──
    elif data in ("ad_all", "ad_registered", "ad_cancel"):
        if data == "ad_cancel":
            await query.edit_message_text("❌ Bekor qilindi.")
            context.user_data.clear()
            return

        context.user_data["ad_target"] = "all" if data == "ad_all" else "registered"
        context.user_data["state"] = "ad_datetime"
        await query.edit_message_text(
            "3️⃣ Yuborish vaqtini kiriting:\n\n"
            "Format: <code>DD.MM.YYYY HH:MM</code>\n"
            "Masalan: <code>25.12.2024 14:30</code>",
            parse_mode="HTML"
        )

    # ── Delete scheduled ad ──
    elif data.startswith("deladid_"):
        ad_id = int(data.split("_")[1])
        await delete_scheduled_ad(ad_id)
        await query.edit_message_text(f"🗑 Reklama #{ad_id} bekor qilindi.")

    # ── Edit text ──
    elif data.startswith("edit_"):
        key = data[5:]
        if key == "cancel":
            await query.edit_message_text("❌ Bekor qilindi.")
            return
        if key in TEXTS:
            context.user_data["edit_key"] = key
            context.user_data["state"] = "edit_waiting"
            current = TEXTS[key][:300]
            await query.edit_message_text(
                f"✏️ Joriy matn:\n\n<code>{current}</code>\n\n"
                f"Yangi matnni yuboring:",
                parse_mode="HTML"
            )


# ─── SCHEDULED ADS WORKER ────────────────────────────────────────────────────

async def ad_worker(bot):
    """Har 60 soniyada rejalashtirilgan reklamalarni tekshiradi"""
    while True:
        await asyncio.sleep(60)
        try:
            ads = await get_pending_ads()
            for ad in ads:
                if ad["target"] == "all":
                    ids = await get_all_start_ids()
                else:
                    ids = await get_all_user_ids()

                ok = 0
                for uid in ids:
                    try:
                        await bot.send_message(chat_id=uid, text=ad["text"])
                        ok += 1
                        await asyncio.sleep(0.05)
                    except Exception:
                        pass

                await mark_ad_sent(ad["id"])
                logger.info(f"📢 Reklama #{ad['id']} yuborildi: {ok} ta")

                # Adminga hisobot
                for admin_id in ADMIN_IDS:
                    try:
                        await bot.send_message(
                            chat_id=admin_id,
                            text=f"📢 Reklama #{ad['id']} avtomatik yuborildi ({ok} ta)"
                        )
                    except Exception:
                        pass
        except Exception as e:
            logger.error(f"Ad worker xatosi: {e}")


# ─── KEEP ALIVE ──────────────────────────────────────────────────────────────

async def keep_alive():
    while True:
        await asyncio.sleep(600)
        try:
            async with httpx.AsyncClient() as client:
                r = await client.get(f"{RENDER_URL}/health", timeout=10)
                logger.info(f"⏰ Keep-alive: {r.status_code}")
        except Exception as e:
            logger.warning(f"⏰ Keep-alive xatosi: {e}")


# ─── BOT SETUP ───────────────────────────────────────────────────────────────

ptb_app = (
    Application.builder()
    .token(BOT_TOKEN)
    .request(HTTPXRequest(
        read_timeout=30, write_timeout=30,
        connect_timeout=30, pool_timeout=30,
    ))
    .build()
)

# Conversation handler (user uchun)
conv = ConversationHandler(
    entry_points=[CommandHandler("start", cmd_start)],
    states={
        REG_NAME:  [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_reg_name)],
        REG_PHONE: [MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), handle_reg_phone)],
        MAIN_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_main_menu)],
    },
    fallbacks=[CommandHandler("start", cmd_start)],
    allow_reentry=True,
)

ptb_app.add_handler(conv)
ptb_app.add_handler(CommandHandler("admin", cmd_admin))

async def admin_or_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin paneli yoki global fallback — lambda o'rniga"""
    if is_admin(update):
        admin_buttons = [
            "📊 Statistika", "👥 Ro'yxat (Excel)",
            "✏️ Matnlarni tahrirlash", "📢 Broadcast",
            "📅 Reklama rejalashtirish", "⏰ Jadval ko'rish", "🔙 Chiqish"
        ]
        if context.user_data.get("state") is None and update.message.text in admin_buttons:
            return await handle_admin_panel(update, context)
        elif context.user_data.get("state") is not None:
            return await handle_broadcast_or_ad_text(update, context)
    return await global_fallback(update, context)

# Admin panel tugmalari + broadcast/ad text qabul
ptb_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, admin_or_fallback))

ptb_app.add_handler(CallbackQueryHandler(handle_callbacks))

# ─── FASTAPI ─────────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    await ptb_app.initialize()
    await ptb_app.bot.set_webhook(url=WEBHOOK_URL)
    logger.info(f"✅ Webhook: {WEBHOOK_URL}")
    asyncio.create_task(keep_alive())
    asyncio.create_task(ad_worker(ptb_app.bot))
    yield
    await ptb_app.shutdown()
    await close_pool()
    logger.info("🛑 Bot to'xtatildi")

web = FastAPI(lifespan=lifespan)

@web.get("/")
@web.get("/health")
async def health():
    return {"status": "ok"}

@web.post(WEBHOOK_PATH)
async def webhook(request: Request):
    data = await request.json()
    update = Update.de_json(data, ptb_app.bot)
    await ptb_app.process_update(update)
    return {"ok": True}
